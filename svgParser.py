from Tools.scripts.make_ctype import values
from openpyxl import Workbook

import xml.etree.ElementTree as ET

from openpyxl.reader.excel import load_workbook

svg_file_path = 'barcode_svg.svg'


tree = ET.parse(svg_file_path)

root = tree.getroot()



# Find all text elements in the SVG

text_elements = root.findall('.//{http://www.w3.org/2000/svg}rect')


# Extract and print the text content
listx = []
listWidth = []
for text_element in text_elements:

    text_content = text_element.get('width')
    text_content2 = text_element.get('x')

    if text_content:
        listWidth.append(text_content)

    if text_content2:
        listx.append(text_content2)

listWidth = listWidth[1:]
for x in range(len(listx)):
    listx[x] = float(listx[x].replace('mm', ''))

for x in range(len(listWidth)):
    listWidth[x] = float(listWidth[x].replace('mm', ''))
firstElemetn = listx[0]
for x in range(len(listx)):
    listx[x] = round(listx[x] - firstElemetn, 2)

listx = [int(round(x/0.2, 0)) for x in listx]
listWidth = [int(round(x/0.2, 0)) for x in listWidth]
print(listx)
print(listWidth)

cels = []

darkCels = []

indeces = []
for x in range(0, listx[-1]+listWidth[-1]):
    indeces.append(x)
for x in range(0, len(listx)):
    darkCelsIndeces = listx[x]+listWidth[x]-1
    for y in range(listx[x], darkCelsIndeces+1):
        darkCels.append(y)
print(darkCels)

for x in range(0, listx[-1]+listWidth[-1]):
    if x in darkCels:
        cels.append(100)
    else:
        cels.append(0)
print(cels)


wb = Workbook()
ws = wb.active

for row in ws.iter_rows(min_row=1, max_col=len(cels), max_row=400):
   iterator = cels.__iter__()
   for cell in row:
        cell.value = iterator.__next__()

#ws.cell(row=2, column=2).value = 2

wb.save("demo.xlsx")


